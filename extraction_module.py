# -*- coding: cp1252 -*-
## don't forget to clip off leading and traling white space.

## Imports----------
from marbles import glass as jc
import ast
import re
import os
RIGHT=1
LEFT=-1

## My workaround encoding issues
#unicode_chars={"—":"-","‘":"'","’":"'","é":"e",'“':'"',"‘":"'"}

## The functions below look useless, but I want it to be easy to flesh them out

## Returns the location of text file containing all utility information (regexes etc.)    
def get_utility_library_directory():
    utility_library_directory="C:\Users\James McGlynn\My Programs\Python Programs\pdf2txt\WorkRelated\Utility_Libraries"
    return utility_library_directory

## Asks the user what library they want to use. It only asks once per time the program runs ATM
def get_current_utility(list_of_utils):
    print "CHOOSE FROM THE LIST BELOW: "
    start=1
    end=len(list_of_utils)
    promptString="MAKE YOUR SELECTION: "
    default=1
    choice=jc.getIntegerInput(start, end, promptString, default,list_of_utils)
    return list_of_utils[choice-1]

## Takes path and returns every other line from the file at that path
## This should be part of a data base.
def get_document_list(text_file):
    with open(text_file,'r') as fh:
        lines=fh.readlines()
    document_list=[]
    for i in range(0,len(lines),2):
        document_list.append(lines[i+1].rstrip('\n'))
        page_list=[]
    return document_list

## Retrieves the info on a desired utility. Returns a dictionary.
def get_utility_library(utility,utility_dictionary_location):
    fh=open(utility_dictionary_location,'r')
    lines_list=fh.readlines()
    whole_doc=''
    for line in lines_list:
        whole_doc=whole_doc+line
    utility_dictionary=ast.literal_eval(whole_doc)
    utility_library_scope_def=utility_dictionary[utility]
    fh.close()
    return utility_library_scope_def

## Returns all available utilities (top level dictionary keys)
def get_utility_list(utility_dictionary_location):
    fh=open(utility_dictionary_location,'r')
    lines_list=fh.readlines()
    whole_doc=''
    for line in lines_list:
        whole_doc=whole_doc+line
    utility_dictionary=ast.literal_eval(whole_doc)
    util_list=[]
    for key in utility_dictionary:
        util_list.append(key)
    fh.close()
    return util_list

## This uses get_index_and_match to find a match for a given
## Regex and then collects some characters in the vicinity.
def get_raw_chars(page_text,library_entry,unicode_chars):
    data_flag=library_entry['data_flag']
    flag_inst=library_entry['data_flag_inst']
    direction=library_entry['direction']
    num_chars=library_entry['raw_char_collect']

    if direction==RIGHT:
        index_and_match=jc.get_index_and_match(page_text,data_flag,flag_inst)
        match_length=len(index_and_match['match'])

        try:
            start_index=index_and_match['index']+match_length
            end_index=start_index+num_chars
            if end_index>len(page_text):
                end_index=len(page_text)
            raw_text=page_text[start_index:end_index].strip()
            
        except:
            start_index=index_and_match['index']
            end_index=index_and_match['index']
            raw_text=index_and_match['index']
      
    elif direction==LEFT:
        ## if left, just find the index of the flag
        index_and_match=jc.get_index_and_match(page_text,data_flag,flag_inst)
        match_length=len(index_and_match['match'])
        
        try:
        ## And subtract to get the right character range
            end_index=index_and_match['index'] - 1
            start_index=end_index-num_chars
            if start_index<0:
                start_index=0
            raw_text=page_text[start_index:end_index].strip()
            
        except:
            start_index=index_and_match['index']
            end_index=index_and_match['index']
            raw_text=index_and_match['index']

    # This was added later to try and prevent encoding errors from popping up.
    # It looks for problem characters and replaces them with more regular ones.
    # The only affected part of the program is printing the raw chars to the xlsx
    raw_text_string_2=""
    matched=0
    for char in raw_text:
        matched=0
        for key in unicode_chars.iterkeys():
            if char==key:
                print char,
                print "equals",
                print key
                raw_text_string_2=raw_text_string_2+unicode_chars[key]
                print raw_text_string2
                matched=1
            else:pass
        if matched==0:
            raw_text_string_2=raw_text_string_2+char
        else:pass
    raw_text=raw_text_string_2    
            
    return {'raw_text':raw_text,'match':index_and_match['match']}

def get_ref_text(raw_text_string,library_entry):
    ## The idea is to take the raw chars and the instructions and get the data
    ## The raw text function gets one peice of info at a time it looks like.
    ## So this should do the same.

    ## Initializations using the library
    collection_method=library_entry['collection_method']
    left_bound_regex=library_entry['left_bound_regex']
    right_bound_regex=library_entry['right_bound_regex']
    data_regex=library_entry['data_regex']
    character_list=library_entry['character_list']
    character_trans=library_entry['character_trans']

    ## If the raw text reads no matches or instance not found, then the refined text is
    ## "No Raw Text"
    if raw_text_string=="NO MATCHES":
        return "No Raw Text"
    elif raw_text_string=="INSTANCE NOT FOUND":
        return "No Raw Text"

    ## If it is not, we continue. What is the collection method?
    ## Meaning - Am I using regular expressions to fine the bounds of the desired information
    ## or the information itself. 
    elif collection_method=='bounds': ##If I'm using the bounds
        if left_bound_regex=="null": ## Check to see what the LB Regex is
            start_index=0 ## If it's the literal string "null" then set start_index to start of string
        else:  ## If it isn't, find the left bound using the leftbound regex
               ## and set the start_index equal to its instance. 
            index_and_match=jc.get_index_and_match(raw_text_string,left_bound_regex,1)

            # Try is necessary here incase the left bound regex doesn't come up with anything
            ## But why isn't there one for the right bound?
            try:
                start_index=index_and_match['index']+len(index_and_match['match'])
            except:
                start_index=index_and_match['index']
        if right_bound_regex=="null":
            end_index=len(raw_text_string)
        else:
            index_and_match=jc.get_index_and_match(raw_text_string,right_bound_regex,1)
            end_index=index_and_match['index']
    elif collection_method=='data':
        index_and_match=jc.get_index_and_match(raw_text_string,data_regex,1)
        start_index=index_and_match['index']
        try:
            end_index=len(index_and_match['match'])+start_index
        except:
            end_index=len(raw_text_string)

    try:        
        ref_text=raw_text_string[start_index:end_index]
    except:
        ref_text="Bad Raw Text"
        
    if ref_text=="Bad Raw Text":
        pass
    elif ref_text=="No Raw Text":
        pass
    else:
        if character_list != "none":
            ref_text=jc.character_selection(ref_text,character_list)
        if character_trans != "none":
            ref_text=jc.character_transform(ref_text,character_trans)

    return ref_text

#### this is taking a string(denoting the util_library), a dictionary (of what is to be printed)
#### a file path to make the xlsx file from, and a flag so you know what type of data you have?
##book_name=print_to_workbook(utility_library,raw_text_dict,ocr_text_path,"raw_text")

def print_to_workbook(data_order_list,dict_to_print,source_path,tab_name):
    
    ## Openpyxl library imports
    from openpyxl import Workbook
    from openpyxl import load_workbook

    ## Get the excel filename from the name of the text file that the user navigated to
    book_name=source_path[:source_path.rindex('.')]+'.xlsx'#jc.getFilenameFromPath(source_path)+'.xlsx'

    try: ## Try opening the workbook with the same name"
        #Found Workbook
        wb=load_workbook(book_name)
        new_wb=-1
        try:
            # Found tab
            ws = wb.get_sheet_by_name(tab_name) #this does not error if it doesn't find what it wants
            if ws==None:
                1+"one"#throw error"
            else:
                pass
        except:
            #Found workbook, but didn't find tab"
            ws=wb.create_sheet(-1,tab_name)
            ws=wb.get_sheet_by_name(tab_name)
            new_wb=-1
            i=0
            for key in data_order_list['library_info']['collection_order']:
                c=ws.cell(row=0, column=i)
                c.value=key
                i=i+1
        
    except: ## If it doesn't exist then start a new workbook in memory
        #Didn't find workbook
        wb = Workbook()
        ws = wb.create_sheet(-1,tab_name)
        ws = wb.get_sheet_by_name(tab_name)
        new_wb=1
        i=0
        for key in data_order_list['library_info']['collection_order']:
            c=ws.cell(row=0, column=i)
            c.value=key
            i=i+1
        
    last_occ_row=ws.rows[-1][0].row
    

    i=0
    for key in data_order_list['library_info']['collection_order']:

        c=ws.cell(row=last_occ_row, column=i)
        try:
            c.value=dict_to_print[key].strip()
        except:
            #Using this area to keep track of encoding errors
            print "This text caused an error"
            print dict_to_print[key]
            c.value="Raw text contained bad chars, see intepreter."
            #c.value=dict_to_print[key]
        i=i+1
            
    wb.save(book_name)
    return book_name

def txt2xlsx(ocr_text_path,utility_library,unicode_chars):

    ## Gets the user to select the document containing the OCR'ed text
    #ocr_text_path=jc.getPath(jc.get_default_directory())[i]

    ## Uses that path to extract the data.
    document_list=get_document_list(ocr_text_path)

    ## Retrieves the utility library from the utility dictionary for the specified utility
    ## which is currently determined with a function (That just returns "Consolidated Edison"

    
    #utility_library=get_utility_library(get_current_utility(),get_utility_library_directory())

    ## Sample of using def get raw chars
    #library_entry_sample=utility_library['extraction_parameters']['G&T Demand1']
    #raw_chars=get_raw_chars(document_list[0],library_entry_sample)

    ## Collecting the raw characters, but I would also like to collect
    ## the flags found by the regular expressions.
    match_dict={}
    raw_text_dict={}
    ref_text_dict={}
    all_results_dict={} ## I might not use this
    for i in range(len(document_list)):
        #print "-----------------------------------------------------------"
        print "Extracting Raw Text for Page: "+str(i+1)
        #print "-----------------------------------------------------------"
        for key in utility_library['library_info']['collection_order']:
            #print utility_library['extraction_parameters'][key]
            raw_text_and_match=get_raw_chars(document_list[i],utility_library['extraction_parameters'][key],unicode_chars)
            raw_text=raw_text_and_match['raw_text']
            match=raw_text_and_match['match']
            #print str(key)+" : "+str(raw_text)+" : "+str(match)

            #This is how the function call would go
            #print "raw_text: "+raw_text
            #print "key: "+key
            ref_text=get_ref_text(raw_text,utility_library['extraction_parameters'][key])
            ref_text_dict[key]=ref_text
            
            raw_text_dict[key]=raw_text
            match_dict[key]=match

        # raw text dict is currently for page of data, not for file of data. 
        
        book_name=print_to_workbook(utility_library,raw_text_dict,ocr_text_path,"raw_text")
        book_name=print_to_workbook(utility_library,ref_text_dict,ocr_text_path,"ref_text")

    return book_name


#book_name=txt2xlsx(0)



