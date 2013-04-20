## This is a stand alone file that can be used after
## the text files are already created. It is not a module

## Declarations for readability
RIGHT=1
LEFT=-1

#imports 
import re

## This is the extraction library - it is particular to a utility. Maybe even more particular than that. 
## This library will eventually go into a text file or database type thing if I'm feeling up to it. 

refined_ConEd_Lib ={'library_info':{ 

                        'utility':'Consolidated Edison',
                        
                        'collection_order':('Name','Account Number','Rate Structure','G&T Demand1',
                                            'G&T Demand2','Primary Demand1','Primary Demand2','Secondary Demand1','Secondary Demand2','Date'),
                        
                        'heading_order':('Name','Account Number','Rate Structure','Date','G&T Demand1',
					'G&T Demand2','Primary Demand1','Primary Demand2','Secondary Demand1','Secondary Demand2')},
                    
#Extraction parameters contains instructions for collecting each peice of data
		    'extraction_parameters':{ 

                        'Name':{
                                'data_flag'                : 'Name:',
                                'data_flag_inst'           : 1,
                                'direction'                : RIGHT,
                                'raw_char_collect'         : 40,
                                'collection_method'        : 'keyword',
                                'collection_parameter_1'   : 'Account',
                                'collection_parameter_2'   : 1,
                                'include_in_final'         : 'NO',
                                'collection_type'          : 0},

                        'Account Number':{
                                'data_flag'                : 'Account number:',
                                'data_flag_inst'           : 1,
                                'direction'                : RIGHT,
                                'raw_char_collect'         : 40,
                                'collection_method'        : 'keyword',
                                'collection_parameter_1'   : "Billing",
                                'collection_parameter_2'   : 1,
                                'include_in_final'         : 'NO',
                                'collection_type'          : 0},

                        'Rate Structure':{
                                'data_flag'                : 'Rate:',
                                'data_flag_inst'           : 1,
                                'direction'                : RIGHT,
                                'raw_char_collect'         : 50,
                                'collection_method'        : 'keyword',
                                'collection_parameter_1'   : 'We me',
                                'collection_parameter_2'   : 1,
                                'include_in_final'         : 'YES',
                                'collection_type'          : 0},
								
                        'Date':{
                                'data_flag'                : 'from',
                                'data_flag_inst'           : 1,
                                'direction'                : RIGHT,
                                'raw_char_collect'         : 40,
                                'collection_method'        : 'keyword',
                                'collection_parameter_1'   : 'Rate',
                                'collection_parameter_2'   : 1,
                                'include_in_final'         : 'YES',
                                'collection_type'          : 0},
								
			'G&T Demand1':{
                                'data_flag'                : """G[^&]{0,2}&[^T]{0,2}T""",
                                'data_flag_inst'           : 1,
                                'direction'                : RIGHT,
                                'raw_char_collect'         : 30,
                                'collection_method'        : 'numbers-with-buffer-decimal-number',
                                'collection_parameter_1'   : 1,
                                'collection_parameter_2'   : 2,
                                'include_in_final'         : 'YES',
                                'collection_type'          : 0},	
								
			'G&T Demand2':{
                                'data_flag'                : """G[^&]{0,2}&[^T]{0,2}T""",
                                'data_flag_inst'           : 1,
                                'direction'                : RIGHT,
                                'raw_char_collect'         : 30,
                                'collection_method'        : 'numbers-with-buffer-decimal-number',
                                'collection_parameter_1'   : 2,
                                'collection_parameter_2'   : 1,
                                'include_in_final'         : 'YES',
                                'collection_type'          : 0},	
								
                        'Primary Demand1':{
                                'data_flag'                : 'Primary demand',
                                'data_flag_inst'           : 1,
                                'direction'                : RIGHT,
                                'raw_char_collect'         : 30,
                                'collection_method'        : 'keyword',
                                'collection_parameter_1'   : 'k',
                                'collection_parameter_2'   : 1,
                                'include_in_final'         : 'YES',
                                'collection_type'          : 0},	
								
                        'Primary Demand2':{
                                'data_flag'                : 'Primary demand',
                                'data_flag_inst'           : 2,
                                'direction'                : RIGHT,
                                'raw_char_collect'         : 30,
                                'collection_method'        : 'keyword',
                                'collection_parameter_1'   : 'k',
                                'collection_parameter_2'   : 1,
                                'include_in_final'         : 'YES',
                                'collection_type'          : 0},

                        'Secondary Demand1':{
                                'data_flag'                : 'Secondary demand',
                                'data_flag_inst'           : 1,
                                'direction'                : RIGHT,
                                'raw_char_collect'         : 30,
                                'collection_method'        : 'keyword',
                                'collection_parameter_1'   : 'kW',
                                'collection_parameter_2'   : 1,
                                'include_in_final'         : 'YES',
                                'collection_type'          : 0},	

                        'Secondary Demand2':{
                                'data_flag'                : 'Secondary demand',
                                'data_flag_inst'           : 2,
                                'direction'                : RIGHT,
                                'raw_char_collect'         : 30,
                                'collection_method'        : 'keyword',
                                'collection_parameter_1'   : 'kW',
                                'collection_parameter_2'   : 1,
                                'include_in_final'         : 'YES',
                                'collection_type'          : 0}
                        }}

##Add this in once the others are working - The idea is to be able to do some basic calculation
## While creating the spreadsheet - I also need to work on casting the data I collect to minimize
## The work I have to do after I open the excel file. 
##,[1,"Cost","ADD","Total-Discount","Total-Discount"])

## Character Lists for Certain types of collection methods - I want to change the names of these.
## They no longer pertain to just one collection method. 
numbers_with_buffer_decimal_number=(',','.','1','2','3','4','5','6','7','8','9','0')
numbers_with_buffer_date=('/','|','1','2','3','4','5','6','7','8','9','0')

## Takes some text and looks for the desired instance of the flag, starting in the direction you choose
## I should add an optional argument to start at a desired index.
def get_index(text,flag,inst,direc):
    if direc==RIGHT:                ## If the direction is the right
        char_index=-1               ## Then you can start at the beginning of the text
        for i in range(inst):       ## Do this once for every instance (Probably a slow way to do this)
            try:                    ## See if you can find the desired string
                #match=re.match(flag,text[(char_index+1):]).group()
                #print match
                char_index=text.index(flag,char_index+1)  ## if you can, get the index

                try:
                    match=re.search(flag,text)
                    print match.group()
                except:
                    print "no match"
     
            except:   
                char_index="Instance Not Found"     ## if you can't, make char_index a string
                
        ## If you're looking for the second instance, the starting char
        ## has now been reset and you look for the string again.
        ## if at any time you don't find the string, it doesn't exit the loop (would be smart)
        ## it just keeps assigning the same string to char_index because it won't be able to
        ## use .index properly. It then returns either a string of failure or an index of success
        return char_index
    
    ## If the direction is LEFT
    ## I couldn't get the starting index of rindex to work properly so it's slightly different
    else:                           ## If the direction is left. 
        char_index=len(text)        ## Set the index to the end of the string
        for i in range(inst):       ## Do this once for every instance
            try:                    ## See if you can find the string (As char-index is reset,
                                    ##    the piece of string being looked at gets smaller)
                char_index=text[:char_index].rindex(flag)
            except:                 ## If you can't find it, make the index a string
                char_index="Instance Not Found"  ## So that all subsequent searches fail. 
        return char_index           

## This is a printer I slapped together to see if I was collecting the right stuff,
## which is good, because I totally wasn't
## I replaced this with an xlsx printer (from openpyxl)
def csv_printer(data_list,file_path):
    file_handle=file(file_path,'a')
    for item in data_list:
        file_handle.write('"'),
        file_handle.write(str(item[1])),
        file_handle.write('"'),
        file_handle.write(",")
    file_handle.write("\n")
    file_handle.close()

## This collects the desired number of characters to be collected for further analysis
## That way I can search for a string and if I don't find it, it doesn't have to go through
## The entire page text and maybe find an instance I don't want or something.
def get_raw_chars(page_text,data_flag,flag_inst,direction,num_chars):
    try:
        if direction==RIGHT:
            ## If going to the right find flag, and make start index right after the last char of the flag
            start_index=get_index(page_text,data_flag,flag_inst,RIGHT)+len(data_flag)
            ## Get end_index be adding 
            end_index=start_index+num_chars
            raw_text=page_text[start_index:end_index].strip()
        elif direction==LEFT:
            ## if left, just find the index of the flag
            end_index=get_index(page_text,data_flag,flag_inst,LEFT)
            ## And subtract to get the right character range
            start_index=end_index-num_chars-len(data_flag)
            raw_text=page_text[start_index:end_index].strip()
    except:
        ## :(
        raw_text="FLAG NOT FOUND"
    return raw_text


## ----------------- extraction functions for every case--------(yeah right)
## This section tries to cover every way you would possible want to go from one string
## to another. by number of characters, by spaces, by a certain string, and it
## also tries to let you collect only the characters you want, numbers, letters, other, combinations

## This takes the a library entry and some text and looks for a specific string
## which it uses as a flag to stop collecting
def stop_key_string(lib_entry,raw_text):
    direction = lib_entry['direction']
    stop_key = lib_entry['collection_parameter_1']
    stop_inst = lib_entry['collection_parameter_2']
    if direction==RIGHT:
        end_index=get_index(raw_text,stop_key,stop_inst,RIGHT)
        if end_index=="Instance Not Found":
            end_index=None
        refined_text=raw_text[:end_index].strip()
    else:
        start_index=get_index(raw_text,stop_key,stop_inst,LEFT)
        if start_index=="Instance Not Found":
            start_index=None
        refined_text=raw_text[start_index:].strip()
    return refined_text

## This just gets a specific number of chars, this will probably only be used if
## I can't figure out any other way to extract something and it happens to be a
## consistent length
def stop_key_number_of_chars(lib_entry,raw_text):
    direction = lib_entry['direction']
    stop_key = lib_entry['collection_parameter_1']
    stop_inst = lib_entry['collection_parameter_2']
    if direction==RIGHT:
        end_index=stop_key
        refined_text=raw_text[:end_index].strip()
    else:
        start_index=len(raw_text)-stop_key
        refined_text=raw_text[start_index:].strip()
    return refined_text

## This collects "words". I can say I want to collect that group of
## characters seperated by spaces on both sides that is x number of
## "words" away from the beginning/end of the raw text string
def stop_key_space_delimited(lib_entry, raw_text):
    direction = lib_entry['direction']
    stop_key = lib_entry['collection_parameter_1']
    stop_inst = lib_entry['collection_parameter_2']
    if direction==RIGHT:
        start_index=get_index(" "+raw_text," ",stop_key,RIGHT)
        end_index=get_index(" "+raw_text," ",stop_inst,RIGHT)
        refined_text=raw_text[start_index:end_index].strip()
    else:
        start_index=get_index(raw_text," ",stop_key,LEFT)
        end_index=get_index(raw_text," ",stop_inst,LEFT)
        refined_text=raw_text[start_index:end_index].strip()  
    return refined_text

## This takes a string and misleadingly, will collect only the characters
## in a character list (most of the time numbers periods and commas)
## with a specified buffer. Meaning that if it fails to find a character in
## the list after x characters it will stop collecting.
def stop_key_numbers_with_buffer(lib_entry, raw_text, character_list):
    direction = lib_entry['direction']
    stop_key = lib_entry['collection_parameter_1']
    stop_inst = lib_entry['collection_parameter_2']
    collection_string=""
    buffer_length=stop_key
    if direction==RIGHT:
        i=0
        while (buffer_length>0 and i<len(raw_text)):
            current_char=raw_text[i]
            acceptable=0
            for acceptable_char in character_list:
                if current_char==acceptable_char:
                    acceptable=1
                else:
                    pass
            if acceptable==1:
                collection_string=collection_string+current_char
                buffer_length=stop_key
            else:
                buffer_length=buffer_length-1
            i=i+1
        refined_text=collection_string.strip()
    else:
        i=len(raw_text)
        while (buffer_length>0 and i>0):
            current_char=raw_text[i-1]
            acceptable=0
            for acceptable_char in character_list:
                if current_char==acceptable_char:
                    acceptable=1
                else:
                    pass
            if acceptable==1:
                collection_string=collection_string+current_char
                buffer_length=stop_key
            else:
                buffer_length=buffer_length-1
            i=i-1
        refined_text=collection_string.strip()
    return refined_text

## Calls the keyword method, and then appends the initial flag text to it
## I wrote this one specifically to look for the words "FUEL OIL" in the
## raw text from the pdf, then find '#4' based on looking for the # sign.
## then I thought #4 was too short so I added it on to make it "#4 FUEL OIL"
## maybe it will be useful for something else one day
def stop_key_keyword_include(lib_entry, raw_text):
    refined_text=stop_key_string(lib_entry,raw_text)+" "+lib_entry['data_flag']#lib_entry[2]
    return refined_text

## This gets a group of chars from the space method and then refines the chars
## based on a character list using the buffer method
## I should really add the refine case as on option to the library so I can do
## it to any refined text as one last refinment step.
## But I was too lazy to do it at the time of this writing.
def stop_key_space_delimited_refine_chars(lib_entry, raw_text, character_list, buffer_length):
    refined_text_inter=stop_key_space_delimited(lib_entry, raw_text)
    import copy
    lib_entry_copy=copy.deepcopy(lib_entry)
    lib_entry_copy['collection_parameter_1']=buffer_length
    lib_entry_copy['collection_parameter_2']=1
    lib_entry_copy['direction']=RIGHT
    refined_text=stop_key_numbers_with_buffer(lib_entry_copy, refined_text_inter, character_list)
    return refined_text

## Default, if there is no function written for the desired case just return the raw text
## as refined text
def stop_key_default(lib_entry, raw_text):
    return raw_text
    
##------------------------End extraction function section------------------
##-------------------------------------------------------------------------

## This function uses the library and the functions above to get the desired data
## found in the vicinity of each keyword. It is a huge if else if statement
def refined_extract(refined_lib,page_text):
    ## Actually a dict
    raw_text_list={}

## The raw text has to be put into a dict
## And the refined text should also be put into a dict.

    ## Let the user know whats coming
    print "----------------------------------------------------------------------"
    print "COLLECTIN/PRINTING RAW TEXT"
    print "----------------------------------------------------------------------"

    #Get the collection order from the utility library and collect text for each item.
    for item in refined_lib['library_info']['collection_order']:
        ## entry will be a dict containing the instruction for item
        entry=refined_lib['extraction_parameters'][item]

        ## A bunch of declarations to pull the data out of the dict
        coll_type=entry['collection_method']
        #data_type=entry['']
        data_flag=entry['data_flag']
        flag_inst=entry['data_flag_inst']
        direction=entry['direction']
        num_chars=entry['raw_char_collect']
        stop_mthd=entry['collection_method']
        stop_key=entry['collection_parameter_1']
        stop_inst=entry['collection_parameter_2']
        fin_tab=entry['collection_type']
        
        ## Call get_raw_chars to get characters in the vicinity of the flag to make
        ## Analysis easier/ more reliable.
        raw_text=get_raw_chars(page_text,data_flag,flag_inst,direction,num_chars)

        ## Add all the results to a raw_text_dict, that way, the raw text can be called on
        ## Using the "Heading order" info from the utility library
        raw_text_list[item]=raw_text

        print item+": "+raw_text_list[item]

    ## This is just to show the user some progress. 
    ##for item in raw_text_list.itervalues(): print item

##    ## Let user know what's coming
##    print "----------------------------------------------------------------------"
##    print "GETTING REFINED DATA"
##    print "----------------------------------------------------------------------"

    refined_text_list={}
    for key in refined_lib['library_info']['collection_order']:

        if refined_lib['extraction_parameters'][key]['collection_method']=="number-of-chars":
            #print "Keyword matched, using number-of-chars method to retrieve data"
            refined_text=stop_key_number_of_chars(refined_lib['extraction_parameters'][key],raw_text_list[key])
            #print refined_text
            refined_text_list[key]=refined_text
            
        elif refined_lib['extraction_parameters'][key]['collection_method']=="keyword":
            #print "Keyword matched, using keyword method to retrieve data"
            refined_text=stop_key_string(refined_lib['extraction_parameters'][key],raw_text_list[key])
            #print refined_text
            refined_text_list[key]=refined_text
            
        elif refined_lib['extraction_parameters'][key]['collection_method']=="space-delimited":
            #print "Keyword matched, using space-delimited method to retrieve data"
            refined_text=stop_key_space_delimited(refined_lib['extraction_parameters'][key], raw_text_list[key])
            #refined_text_list.append(refined_text)
            refined_text_list[key]=refined_text

        elif refined_lib['extraction_parameters'][key]['collection_method']=="numbers-with-buffer-decimal-number":
            #print "Keyword matched, using numbers-with-buffer-decimal-number method to retrieve data"
            refined_text=stop_key_numbers_with_buffer(refined_lib['extraction_parameters'][key], raw_text_list[key],numbers_with_buffer_decimal_number)
            #refined_text_list.append(refined_text)
            refined_text_list[key]=refined_text
            
        elif refined_lib['extraction_parameters'][key]['collection_method']=="numbers-with-buffer-date":
            #print "Keyword matched, using numbers-with-buffer-date method to retrieve data"
            refined_text=stop_key_numbers_with_buffer(refined_lib['extraction_parameters'][key], raw_text_list[key],numbers_with_buffer_date)
            #refined_text_list.append(refined_text)
            refined_text_list[key]=refined_text
            
        elif refined_lib['extraction_parameters'][key]['collection_method']=="keyword-include":
            #print "Keyword matched, using keyword-include method to retrieve data"
            refined_text=stop_key_keyword_include(refined_lib['extraction_parameters'][key], raw_text_list[key])
            #refined_text_list.append(refined_text)
            refined_text_list[key]=refined_text

        elif refined_lib['extraction_parameters'][key]['collection_method']=="space-delimited-refine-chars":
            #print "Keyword matched, using space-delimited-refine-chars method to retrieve data"
            refined_text=stop_key_space_delimited_refine_chars(refined_lib['extraction_parameters'][key], raw_text_list[key], numbers_with_buffer_decimal_number, 3)
            #refined_text_list.append(refined_text)
            refined_text_list[key]=refined_text
            
        else:
            print "Keyword did not match any function, '"+refined_lib['extraction_parameters'][key]['collection_method']+"'"
            #refined_text_list.append(raw_text_list[i])
            refined_text_list[key]=refined_text

    ##Let user know what's coming
    print "----------------------------------------------------------------------"
    print "COLLECTING/PRINTING REFINED TEXT"
    print "----------------------------------------------------------------------"
    for key in refined_lib['library_info']['heading_order']:
        print key+": "+refined_text_list[key]

    return refined_text_list

def get_document_list(doc_list):
    page_list=[]
    document_list=[]
    for i in range(0,len(lines),2):
        page_list.append(lines[i].rstrip('\n'))
        page_list.append(lines[i+1].rstrip('\n'))
        document_list.append(page_list)
        page_list=[]
    return document_list

def extract_information(library,document_list):
    refined_results=[]
    for i in range(len(document_list)):
        raw_text=document_list[i][1]
        result=refined_extract(library, raw_text)   
        refined_results.append(result)
    return refined_results

###--------------------------------------------------------------------------------------------
###--------START OF PROGRAM---------------------------------------------------------------
import os
import useful

## Testing directory
default_directory="C:\Users\James McGlynn\My Programs\Python Programs\pdf2txt\WorkRelated"

## Choose library here
library=refined_ConEd_Lib

## Have user navigate to text file
text_file=useful.getPath(default_directory)

## Open the text file and put every line in a list
with open(text_file,'r') as fh:
    lines=fh.readlines()
    
## Pull every other line from the list (That's just how I made the document)
document_list=get_document_list(lines)

## Use the library and the text to try and get the data you want
refined_results=extract_information(library, document_list)


def print_by_account_number(library,refined_results,text_file):

    ## Openpyxl library imports
    from openpyxl import Workbook
    from openpyxl import load_workbook

    ## Get the excel filename from the name of the text file that the user navigated to
    book_name=useful.getFilenameFromPath(text_file)+'.xlsx'

    try: ## Try opening the workbook with the same name"
        wb=load_workbook(book_name)
    except: ## If it doesn't exist then start a new workbook in memory
        wb = Workbook()

    ## This match business is to try and sort the gathered data by account number, but the account numbers
    ## aren't really reliable enough, so I think I'm going to stop doing this for now. 
    match=0
    for entry in refined_results:
        
        sheet_names=wb.get_sheet_names()
        for item in sheet_names:
            if entry['Account Number']==item:
                match=1
            else:
                pass
        if match==1:
            ws=wb.get_sheet_by_name(entry['Account Number'])
            last_occ_row=ws.rows[-1][0].row
            i=0
            for key in library['library_info']['heading_order']:
                c=ws.cell(row=last_occ_row, column=i)
                c.value=entry[key]
                i=i+1
        else:
            newSheet=wb.create_sheet()
            newSheet.title = str(entry['Account Number'])
            i=0
            for key in library['library_info']['heading_order']:
                c=newSheet.cell(row=0, column=i)
                d=newSheet.cell(row=1, column=i)
                c.value=key
                d.value=entry[key]
                i=i+1
        match=0
            
    wb.save(book_name)
    return book_name

def print_by_none(library,refined_results,text_file):

    ## Openpyxl library imports
    from openpyxl import Workbook
    from openpyxl import load_workbook

    ## Get the excel filename from the name of the text file that the user navigated to
    book_name=useful.getFilenameFromPath(text_file)+'.xlsx'

    try: ## Try opening the workbook with the same name"
        wb=load_workbook(book_name)
        new_wb=-1
        ws = wb.get_active_sheet()
        
    except: ## If it doesn't exist then start a new workbook in memory
        wb = Workbook()
        ws = wb.get_active_sheet()
        new_wb=1
        i=0
        for key in library['library_info']['heading_order']:
            c=ws.cell(row=0, column=i)
            c.value=key
            i=i+1
    
    for entry in refined_results:
    
        last_occ_row=ws.rows[-1][0].row

        i=0
        for key in library['library_info']['heading_order']:
            c=ws.cell(row=last_occ_row, column=i)
            c.value=entry[key]
            i=i+1
            
    wb.save(book_name)
    return book_name


excel_output_file=print_by_none(library, refined_results, text_file)
print 'done!'









