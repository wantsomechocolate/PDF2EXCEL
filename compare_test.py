## TESTING GROUND FOR COMPARING TEXT BASED IMAGES

print "Importing necessary libraries"

## IMPORT COMPARE (WRITTEN BY ME)
import compare

## GET THE PYTHON IMAGE LIBRARY UP IN HERE
from PIL import Image as pili

## IMPORT OS MAINLY TO GET FILENAMES FROM PATHS
import os

## IMPORT OPEN CV FOR IMAGE ARRAY MANIPULATION
import cv2

## IMPORT NUMPY BECAUSE OPEN CV USES NUMPY ARRAYS FOR STORING IMAGE DATA
import numpy as np

## IMPORT A "PACKAGE" I MADE TO STORE SOME USEFUL FUNCTIONS IN
from marbles import glass as useful

## IMPORT TIME SO I KNOW HOW LONG STUFF TAKES
from time import time

## IMPORT PYXL TO CREATE XLSX (FOR TESTING RESULTS)
from openpyxl import Workbook
from openpyxl import load_workbook

print "Initializing read and write locations"

## TIMESTAMP TO USE TO MAKE RESULTS UNIQUE
iteration_time_stamp=str(int(time()))+'/'

## BASE TESTING DIRECTORY
base_dir="C:/Users/James McGlynn/My Programs/Python Programs/pdf2txt/CompareExperiment/compare/images/"

## DIRECTORY TO RETRIEVE TEST IMAGES FROM
wdir=base_dir+"images_set_to_test/"

#DIRECTORY TO SAVE MANIPULATED IMAGES TO
save_dir=base_dir+"images_manipulated/"+iteration_time_stamp

#DIRECTORY TO SAVE SPREADSHEET OF COMPARE RESULTS (IM1,IM2,PERCENT SIMILAR)
compare_dir=base_dir+ "compare_results/" + iteration_time_stamp

#DIRECTORY TO SAVE IMAGES ONCE THEY HAVE BEEN GROUPED FOR EASY VISUAL CHECK OF PROGRAM SUCCESS
images_groups_dir=base_dir+"images_groups/" + iteration_time_stamp

print "Retrieving and sorting previously generated set of supposedly rotated and cropped images"

## GET A LIST OF PNG IMAGES FROM THE GIVEN DIRECTORY
image_handle_list=useful.image_handles_from_dir_png(wdir)

## SORT THE LIST BECAUSE IT HELPS ME ANALYZE RESULTS
# Make space for sorted filenames
sorted_list=[]

# Strip filenames from the image handles to sort based on alphabet
for item in image_handle_list:
    sorted_list.append(os.path.basename(item[0].filename))

# Standard sort
sorted_list.sort()

# I don't understand why this works, but it takes the image_handle_list (list of image handles)
# and moves its contents around to match the order of the the sorted list (list of strings)
for item in image_handle_list:
    image_handle_list[sorted_list.index(os.path.basename(item[0].filename))]=item

print "Preparing images for pixel by pixel comparison via blurring and resizing"

## THESE ARE SOME INITILIZATIONS FOR BLUR. COUNT 1 IS ASKING IF YOU WANT TO BLUR BEFORE SHRINKING THE IMAGE
## COUNT 2 IS ASKING ABOUT AFTER. BEFORE IS OFF BECAUSE IT TAKES A LONG TIME TO BLUR A 3000x2000 PIXEL IMAGE
blur_count1=0
blur_count2=1

## KEEPING TRACK OF TIME
time1=time()

## THIS IS WHERE THE IMAGES ARE PREPARED FOR COMPARISON
## THE IMAGES CAN BE BLURRED AND SHRUNKEN A NUMBER OF TIMES
for i in range(len(image_handle_list)):

    ## BLUR IMAGE IF COUNT1 IS NOT 1

    im_prepared=image_handle_list[i][0]

    if blur_count1>0:
        im_prepared=compare.blurify(im_prepared,1)
    
    else:
        pass
        
    ## RESIZE IMAGE TO 256x256 - IF THIS IS NOT DONE, THAN THERE ARE PROBLEMS DOWNSTREAM WHEN COMPARING THE
    ## IMAGES BECAUSE THEY WON'T BE THE SAME SIZE. I CAN EITHER ALLOW FOR THAT DOWNSTREAM, OR NOT ALLOW IT HERE.   
    im_prepared=compare.resizify(im_prepared,(256,256))
    
    ## BLUR IT AGAIN IF COUNT 2 IS GREATER THAN 0 ,aka 1
    if blur_count2>0:
        im_prepared=compare.blurify(im_prepared,1)

    else:
        pass
        
    ## RESIZE AGAIN......? I CONSIDERED FOR A WHILE TO RESIZE AGAIN TO 16X16, BUT THE MARGIN BETWEEN SUCCESS AND FAILURE
    ## WAS PRETTY TIGHT. LEFT IN FOR CONSIDERATION OF OTHERS.   
    #im_prepared=compare.resizify(im_prepared,(16,16))
    
    ## ADD EDITED IMAGE TO HANDLE LIST. THIS PRODUCES A LIST OF THE SAME LENGTH BUT WITH
    ## TWO ITEMS PER ITEM IN THE LIST (IMAGE AND EDITED IMAGE)
    image_handle_list[i].append(im_prepared)

    ## GIVE VISUAL FEEDBACK OF PROGRESS BY PRINTING DOTS!
    print ".",
    
print ""

## SAVE PREPARED IMAGES IN THE SPECIFIED DIRECTORY

if not os.path.exists(save_dir): os.makedirs(save_dir)
for i in range(len(image_handle_list)):
    ## USE SAME IMAGE NAME AS SPECIFIED BY IMAGE 
    image_handle_list[i][1].save(save_dir+os.path.basename(image_handle_list[i][0].filename))

## KEEP TRACK OF TIME (OF HOW LONG IT TOOK TO PREPARE THE IMAGES)
time2=time()
print "Preparing all images took "+str(int(time2-time1))+" seconds."

print "Comparing images pixel by pixel to get a percent similarity"

## THE NEXT SECTION USES THE PREPARED IMAGES AND COMPARES THEM PIXEL BY PIXEL
image_comparison_list=[]
## FOR EVERY HANDLE
## COMPARE IT TO ALL HANDLES IN THE LIST AT A HIGHER INDEX
## (THIS ENSURES THAT ALL COMPARISONS ARE MADE ONCE
## AND THAT THEY ARE ONLY MADE ONCE.
for i in range(len(image_handle_list)):
    for j in range(len(image_handle_list)):
        if i<j:
            ## Pixel compare is supposed to compare each pixel, which should be one value (grayscale or b&w)
            ## and then divide by the possible max of the procedure. Which for a 256x256 image should be
            ## 256 x 256 (dimensions of image or number of comparisons) multiplied by the max
            ## difference at each pixel which would be white vs black or 255-0 or 255.
            ## The result of the above will be a percent different - subtract by one to get percent sim.
            percent_sim=compare.pixel_compare(image_handle_list[i][1],image_handle_list[j][1])
            image_comparison_list.append([image_handle_list[i][0],image_handle_list[j][0],percent_sim])

## KEEPING TRACK OF TIME
time3=time()
print "Comparing all images took "+str(int(time3-time2))+" seconds."

print "Printing results to an excel file"

## PRINT RESULTS TO XLSX FOR EXAMINATION AND FINE TUNING OF PARAMS

book_name=compare_dir+'compare results.xlsx'

## INITIALIZE WORKBOOK
wb = Workbook()

## GET ACTIVE SHEET (ALL WORKBOOKS ARE CREATED WITH ONE SHEET)
ws=wb.get_active_sheet()

## CREATE A HEADER ROW
c=ws.cell(row=0, column=0)
c.value="Image One"
c=ws.cell(row=0, column=1)
c.value="Image Two"
c=ws.cell(row=0, column=2)
c.value="Similarity"

## FOR EVERY COMPARISON THAT WAS MADE, PRINT THE RESULTS TO A ROW
for i in range(len(image_comparison_list)):

    if i%10==0:
        print ".",
    else: pass
    
    c=ws.cell(row=i+1, column=0)
    c.value=os.path.basename(image_comparison_list[i][0].filename)
    c=ws.cell(row=i+1, column=1)
    c.value=os.path.basename(image_comparison_list[i][1].filename)
    c=ws.cell(row=i+1, column=2)
    c.value=image_comparison_list[i][2]

print ""

## SAVE WORKBOOK
if not os.path.exists(compare_dir): os.makedirs(compare_dir)
wb.save(book_name)

print "Sorting the compared images into groups of similar images"

## ---------START GROUPS------------------------------------------------------------------------

## THIS SECTION ATTEMPTS TO GROUP THE IMAGES BASED ON THE SIMILARITY COMPARISONS. THE GENERAL IDEA IS
## TO TAKE EACH IMAGE, AND MAKE A SET OF ALL THE IMAGES THAT MATCH IT, WHICH PRODUCES AS MANY SETS AS IMAGES
## THEN IT GOES THROUGH ALL THE SETS AND COMBINES ALL SUPER AND SUBSETS (please tell me how to make that
## process better). AFTER THAT PROCESS, I CHECK TO SEE WHAT IMAGES DIDN'T MAKE IT (THESE WILL BE IMAGES
## THAT HAD NO MATCHES) AND PUT THOSE IMAGES IN THE REJECT GROUP. I STILL NEED TO DEAL WITH PARTIALLY UNIQUE
## SETS, I'M THINKING IF 50% OF SET IS SUBSET OF ANOTHER SET, THEN UNION ENTIRE SET, IF LESS THAN 50%
## THEN REMOVE UNIQUE ITEMS AND ADD THEM TO THE REJECT GROUP (ALWAYS THE LAST GROUP)

## IMAGE COMPARISON SUCCESS IS BASED ON THE NUMBER BELOW - IT SHOULD BE PROJECT WIDE
threshold_similarity=0.98

## MAKE SPACE FOR THE LISTS OF IMAGES
similar_images_by_image=[]

## FOR EVERY IMAGE IN THE IMAGE_HANDLE_LIST
for i in range(len(image_handle_list)):

    ## GET A LIST OF IMAGES THAT MATCHED IT FROM COMPARE_RESULTS AND APPEND THAT LIST TO SIMILAR_IMAGES_BY_IMAGE
    similar_images_by_image.append(compare.compare_results_one_image_return_set(image_comparison_list,os.path.basename(image_handle_list[i][0].filename),threshold_similarity))

## DON'T DISTURB THE ORIGINAL LIST OF SIMILAR IMAGES
new_groups_from=list(similar_images_by_image)

## REMOVING DUPLICATE SETS IS AN ITERATIVE PROCESS (AT LEAST 2 REQUIRED FOR CURRENT DATA SET)
## SO I AM PREPARING FOR THAT 
new_groups_to=new_groups_from[:1]

## SOME STUFF TO MAKE THE WHILE LOOP WORK
need_to_remove_duplicate_sets=True
duplicate_sets_found=True

## WHILE THE NEED TO REMOVE DUPLICATE SETS REMAINS
while need_to_remove_duplicate_sets==True:

    ## CHECK TO MAKE SURE THERE ARE DUPLICATE SETS BY ITERATING THROUGH 
    for i in range(len(new_groups_from)):

        ## WITH TWO COUNTERS 
        for j in range(len(new_groups_from)):

            ## THIS IS SO I DON'T WASTE TIME CHECKING THE SAME CASE
            if i<j:

                ## ANY GROUP IS EITHER SUB OR SUPER SET OF ANOTHER, THE PROCESS OF REMOVING DUPLICATE SETS NEEDS TO BE REPEATED
                if new_groups_from[i].issubset(new_groups_from[j]) or new_groups_from[i].issuperset(new_groups_from[j]):

                    ## SO SET THIS TO TRUE
                    duplicate_sets_found=True

    ## CHECK TO SEE IF THE ABOVE PROCESS PRODUCED YES DUPES OR NO DUPES
    ## I THINK THIS IS SUPERFLUOUS, IT IS SOMETHING I PUT IN WHILE DEBUGGING THAT
    ## I DIDN'T NEED TO, I'LL REMOVE IT SOMEDAY.
    if duplicate_sets_found==True:

        ## IF A DUPE WAS FOUND, THEN WE NEED TO REMOVE IT!
        need_to_remove_duplicate_sets=True

    ## IF A DUPE WAS NOT FOUND
    else:

        ## CHANGE FLAG TO FALSE AND SKIP THE DUPE FINDING PROCESS
        need_to_remove_duplicate_sets=False

    ## RESET THE DUPLICATES_SETS_FOUND VARIABLE - THIS MEANS THE WHILE LOOP WILL STILL
    ## HAVE ANOTHER GO, BUT THE PROCESS TO CHECK FOR DUPES MIGHT FAIL AND IT WILL EXIT ON THE NEXT TIME.
    duplicate_sets_found=False

    ## IF AFTER ALL OF THE ABOVE, THE NEED TO REMOVE DUPES ENDS UP MATERIALIZING
    if need_to_remove_duplicate_sets == True:

        ## ITERATE THROUGH THE LIST CONTAINING GROUPS(SETS) OF IMAGES(IMAGE HANDLES)
        for i in range(len(new_groups_from)): #iterate through the list of image groups

            ## RESET FLAG
            match=False

            ## ITERATE THROUGH THE CURRENTLY FOUND DISTINCT GROUPS (IT IS SET UP SO THAT ON THE FIRST PASS,
            ## NEW_GROUPS_TO WILL HAVE 1 ITEM, WHICH WILL BE THE SAME AS THE FIRST ITEM FROM NEW_GROUPS_FROM
            for j in range(len(new_groups_to)): #iterate through new groups

                ## IF MATCH IS FALSE - MEANING IF THE SET BEING LOOKED AT HAS NOT BEEN A SUB OR SUPER SET
                ## OF ANY OTHER SETS
                if match==False:

                    ## THEN KEEP CHECKING FOR SUB SUPER SET NESS - IF YOU FIND IT.
                    if (new_groups_from[i].issubset(new_groups_to[j])) or (new_groups_from[i].issuperset(new_groups_to[j])):

                        ## RECORD THE INDEX OF NEW_GROUPS_FROM
                        index_i=i

                        ## RECORD THE INDEX OF NEW_GROUPS_TO
                        index_j=j

                        ## CHANGE MATCH TO TRUE - THE SET WAS A SUB OR SUPER SET OF ANOTHER SET
                        match=True

                    ## IF THE SET WASN'T A MATCH, JUST GO CHECK THE NEXT ONE.
                    else: pass

            ## NOW THAT WE ARE OUTSIDE OF THE ABOVE FOR LOOP, MATCH WILL EITHER HAVE BEEN
            ## SET TO TRUE IT WILL NOT HAVE BEEN. CHECK TO SEE WHAT THE CASE IS.
            if match==True:

                ## IF TRUE, THEN UNION THE SET WITH WHATEVER SET WITH WHICH IT "MATCHED" (WAS A SUPER OR SUB SET OF)
                new_groups_to[index_j]=new_groups_from[index_i].union(new_groups_to[index_j])

            ## IF MATCH IS FALSE - THE SET WAS UNIQUE
            elif match==False:

                ## ADD IT TO THE LIST OF SETS TO BE INCLUDED FOR COMPARISON ON THE NEXT ROUND (THIS PROCESS IS DONE
                ## ONCE FOR EACH GROUP IN THE ORIGINAL LIST OF SETS.
                new_groups_to.append(new_groups_from[i])

        ## NOW THAT WE ARE OUT OF THE FOR LOOP WHICH ITERATES THROUGH THE ORIGINAL LIST OF SETS
        ## LETS REARRANGE SOME STUFF
        ## SET THE STARTING LIST OF SETS TO THE NEWLY ARRIVED AT LIST OF SETS THAT SHOULD HAVE NONE, OR AT LEASE LESS DUPLICATES
        new_groups_from=list(new_groups_to)

        ## SET ANOTHER VARIABLE TO THE BE EQUAL TO THE SUPPOSEDLY DUPLICATE-LESS LIST OF SETS.
        final_product=list(new_groups_to)

        ## SET NEW_GROUPS_TO EQUAL TO THE FIRST SET IN NEW_GROUPS_FROM, AS IT WILL OBVIOUSLY NOT BE A SUB OR SUPER SET OF NOTHING
        new_groups_to=new_groups_from[:1]

## PRINT OUT SOME USEFUL INFORMATION ABOUT WHAT HAPPENED
print "I found "+str(len(final_product))+" groups."
for i in range(len(final_product)):
    print "Group "+str(i)+" has "+str(len(final_product[i]))+" items."


## IN BETWEEN ABOVE AND BELOW I NEED TO DO ONE MORE THING WHICH IS TO ADDRESS SETS THAT HAVE ITEMS BOTH IN OTHER SETS,
## AND NOT IN ANY OTHER SETS. I CALL THEM PARTIALLY UNIQUE SETS


## AND NOW ANOTHER STEP - THE ABOVE PROCESS DOES NOT GAURUNTEE THAT ALL IMAGES WILL MAKE IT INTO AT LEAST ONE GROUP
unique_images=final_product[0]

## SET UNIQUE_IMAGES EQUAL TO THE UNION OF ALL SETS IN FINAL_PRODUCT
for i in range(len(final_product)):
    unique_images=unique_images.union(final_product[i])

## GET THE LIST WHICH WAS ORIGINALLY INTENDED TO GO THROUGH THIS WHOLE PROCESS OF COMPARING AND GROUPING AND ALL THAT
initial_list=[]
for item in image_handle_list:
    initial_list.append(item[0])
initial_set=set(initial_list)

## USING BUILT IN SET MATH, SUBTRACT THE TWO SETS.
diff_set=initial_set.difference(unique_images)

print "The image(s) below are not in any group because they did not match any other images"


for item in diff_set:
    print os.path.basename(item.filename)

reject_set=diff_set

## THE IMAGES IN THIS SO CALLED REJECT SET WILL BE IMAGES THAT HAD NO MATCHES DURING THE IMAGE COMPARISON PHASE
## I AM CURRENTLY ADDING THEM TO AN ADDITIONAL GROUP THAT THE USER CAN SORT THEMSELVES
final_product_with_rejects=final_product
final_product_with_rejects.append(reject_set)

print "With the reject set there are " +str(len(final_product_with_rejects))+" sets."

print "And I still have to do something about partially unique sets"

## FOR EASY PRINTING OF SOME STUFF
##for sett in final_product:
##	print "group"
##	for item in sett:
##		print os.path.basename(item.filename)

                  
    
#################### INITIALIZATIONS
##################groups=[]
##################group_number=0
##################match=0
##################
#################### ITERATE THOUGH THE COMPARISON LIST
##################for item in image_comparison_list:
##################
##################    ## RESET MATCH
##################    match=0
##################    
##################    ## EVERYTIME A COMPARISON MEETS THE REQUIREMENTS FOR SUCESS
##################    if item[2]>=threshold_similarity:
##################
##################        print ".",
##################
##################        ## START LOOKING THROUGH ALL THE GROUPS CREATED SO FAR
##################        for i in range(len(groups)):
##################
##################            ## IF EITHER IMAGE FROM THE COMPARISON MATCHES ANY IMAGE IN ANY GROUP AND A MATCH HAS NOT YET BEEN FOUND
##################            if ((item[0] in groups[i]) or (item[1] in groups[i])) and match==0:
##################
##################                ## RECORD THE CURRENT GROUP NUMBER
##################                matched_group = i
##################
##################                ## STOP LOOKING FOR A MATCH
##################                match=1
##################
##################        ## IF A MATCH WAS NOT FOUND ABOVE
##################        if match==0:
##################
##################            ## ADD BOTH IMAGES TO A NEW GROUP VIA APPEND
##################            groups.append([item[0], item[1]])
##################        else:
##################            ## IF A MATCH WAS FOUND THEN ADD BOTH IMAGES TO THE GROUP WHERE THE MATCH WAS. 
##################            groups[matched_group].append(item[0])
##################            groups[matched_group].append(item[1])
##################
##################print ""
##################
##################print "Attempting to remove duplicate sets, if any"
##################
####################------------------------------------------------------------------------------------------
#################### THE ABOVE PROCEDURE IS FLAWED, IT LEAVES OUT IMAGES WITH NO MATCHES AND IT ALSO CREATES
#################### GROUPS THAT ARE SUBSETS OF OTHER GROUPS, OR CONTAIN ONLY SUBSETS OF OTHER GROUPS AND IMAGES WITH 
#################### MATCHES NOT PRESENT IN OTHER GROUPS. THE ABOVE BY ITSELF ALSO LEAVES ALL DUPLICATES INCLUDED.
####################-------------------------------------------------------------------------------------------
##################
#################### TO ELIMINTATE DUPLICATES AND TO REMOVE STRAIGHT UP SUBSET GROUPS, THE FOLLOWING IS DONE.      
##################new_groups=[]
##################issub=False
##################
#################### FOR EACH GROUP
##################for i in range(len(groups)):
##################
##################    print ".",
##################
##################    ## CONVERT EACH GROUP TO A SET (A SET BY DEFINITION DOES NOT CONTAIN DUPES (not an original thought))
##################    new_group=list(set(groups[i]))
##################    ## SORT IS FOR MY BENEFIT, I BELIEVE.
##################    new_group.sort()
##################    ## FOR EACH "NEW GROUP" (At first there won't be any)
##################    for i in range(len(new_groups)):
##################        ## IF NEW GROUP (the sorted set) IS A SUBSET OF ANY OF THE NEW GROUPS (previously sorted sets basically)
##################        if set(new_group).issubset(set(new_groups[i])):
##################            ## CHANGE ISSUB TO TRUE
##################            issub=True
##################        ## OTHERWISE DO NOTHING
##################        else: pass
##################
##################    ## IF THE SET WAS IN FACT A SUBSET
##################    if issub==True:
##################        ## DO NOT ADD IT TO THE LIST OF GROUPS
##################        pass
##################
##################    ## OTHERWISE
##################    else:
##################
##################        ## ADD THE SORTED SET TO NEW GROUPS
##################        new_groups.append(new_group)
##################
##################    ## CHANGE ISSUB BACK TO FALSE
##################    issub=False
##################
##################print ""
##################    
####################------------------------------------------------------------------------------------------
#################### THE ABOVE PROCEDURE IS ALSO FLAWED, BECAUSE IT DOESN'T TAKE CARE OF THE CASE WHEN A
#################### SET CONTAINS ONLY ITEMS FOUND IN OTHERS SETS, AND ITEMS THAT ARE UNIQUE TO THAT SET,
#################### IN WHICH CASE I WOULD WANT TO JOIN THOSE TWO SETS (UNDER CERTAIN CRITERIA). FOR EXAMPLE
#################### IF THERE WAS TWO SETS, ONE THAT CONTAINED ALL DUPES AND ONE UNIQUE ITEM, I THINK IT WOULD BE
#################### WISE TO UNION THE GROUPS, BUT IF THE ONE SET CONTAINS ALL UNIQUE ITEMS EXCEPT FOR ONE DUPE, THEN I
#################### WOULD NOT WANT TO JOIN THE SETS BECAUSE THE DUPE WAS LIKELY A MISMATCH. BUT WHERE IS
#################### THE CUTOFF?
##################
#################### THIS METHOD ALSO CURRENTLY DOES NOT CATCH THE CASE WHEN A SUBSET OCCURS BEFORE ITS CONTAINING SET
#################### IN THE LIST OF GROUPS. BUT I'M NOT SURE IF THAT WOULD EVER HAPPEN OR NOT. 
####################-------------------------------------------------------------------------------------------
##################
##################
##################
print "Saving "+ str(len(final_product_with_rejects)) +" groups of images into their own directories"

## KEEPING TRACK OF TIME
time9=time()

## THIS SECTION IS PRETTY SLOW :(
## THE CODE BELOW SAVES EACH GROUP TO ITS OWN DIRECTORY
for i in range(len(final_product_with_rejects)):
    print "Saving "+str(len(final_product_with_rejects[i]))+" items for group "+str(i+1)
    
    group_num="group_"+str(i+1)+'/'

    for item in final_product_with_rejects[i]:

        print ".",

        save_to_dir=images_groups_dir+group_num

        if not os.path.exists(save_to_dir): os.makedirs(save_to_dir)
        item.save(save_to_dir+os.path.basename(item.filename))
        
    print ""

## KEEPING TRACK OF TIME
time10=time()
print "Saving all images took "+str(int(time10-time9))+ " seconds"
#################### --------END GROUPS------------------------------------------------------------------------------

print "Exiting program"
